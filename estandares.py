import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
import openpyxl

# Ruta al archivo Excel
archivo_excel = "estandares.xlsx"

# Cargar el archivo de Excel
def cargar_hojas():
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        hojas_disponibles = wb.sheetnames

        if "Hoja2" not in hojas_disponibles or "Hoja1" not in hojas_disponibles:
            messagebox.showerror("Error", "Las hojas esperadas no existen en el archivo.")
            return None, None, None

        hoja_estandares = wb["Hoja2"]
        hoja_registros = wb["Hoja1"]
        return wb, hoja_estandares, hoja_registros
    except FileNotFoundError:
        messagebox.showerror("Error", "El archivo Excel no se encuentra.")
        return None, None, None

# Buscar estándar por QR en Hoja2
def buscar_estandar(qr_code, hoja_estandares):
    for fila in hoja_estandares.iter_rows(min_row=2, max_row=hoja_estandares.max_row):
        if fila[0].value and str(fila[0].value) == qr_code:
            return fila
    return None

# Actualizar lista de registros en la interfaz
def actualizar_lista(tabla):
    for i in tabla.get_children():
        tabla.delete(i)
    
    wb, hoja_estandares, hoja_registros = cargar_hojas()
    if not wb or not hoja_estandares or not hoja_registros:
        return
    
    for fila in hoja_registros.iter_rows(min_row=2, max_row=hoja_registros.max_row):
        if fila[1].value and fila[4].value is None:  # Mostrar solo los no entregados
            id_producto = fila[0].value
            codigo_producto = fila[1].value
            fecha_retiro = fila[2].value
            empleado_retiro = fila[3].value
            tabla.insert('', 'end', values=(id_producto, codigo_producto, fecha_retiro, empleado_retiro))

# Función al escanear estándar
def escanear_estandar():
    qr_code = qr_entry.get().strip()
    empleado = empleado_entry.get().strip()
    if not qr_code or not empleado:
        messagebox.showwarning("Advertencia", "Escanea un código QR y escribe la rúbrica del empleado.")
        return

    wb, hoja_estandares, hoja_registros = cargar_hojas()
    if not wb or not hoja_estandares or not hoja_registros:
        return

    fila_estandar = buscar_estandar(qr_code, hoja_estandares)
    if not fila_estandar:
        messagebox.showerror("Error", "Estándar no encontrado en el inventario.")
        return

    producto = fila_estandar[1].value
    codigo = fila_estandar[0].value
    for fila in hoja_registros.iter_rows(min_row=2, max_row=hoja_registros.max_row):
        if fila[0].value == producto and fila[4].value is None:
            fila[4].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            fila[5].value = empleado
            estado_label.config(text="Estándar entregado", fg="green")
            break
    else:
        hoja_registros.append([producto, codigo, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), empleado, None, None])
        estado_label.config(text="Estándar retirado por Quimico", fg="green")
    
    try:
        wb.save(archivo_excel)
    except PermissionError:
        messagebox.showerror("Error", "No se pudo guardar el archivo. Cierra Excel e intenta de nuevo.")
        return

    qr_entry.delete(0, tk.END)
    empleado_entry.delete(0, tk.END)
    actualizar_lista(lista_estandares)

# Configurar ventana principal
ventana = tk.Tk()
ventana.title("Sistema de Gestión de Estándares Laboratorio Desarrollo Analitico")
ventana.geometry("1200x768")

# Etiquetas y campos de texto
tk.Label(ventana, text="Escanea el estándar:", font=("Arial", 16)).pack(pady=20)
qr_entry = tk.Entry(ventana, width=50, font=("Arial", 14))
qr_entry.pack(pady=10)

tk.Label(ventana, text="Rúbrica de empleado:", font=("Arial", 16)).pack(pady=20)
empleado_entry = tk.Entry(ventana, width=50, font=("Arial", 14))
empleado_entry.pack(pady=10)

# Botón para escanear y registrar
tk.Button(ventana, text="Registrar", font=("Arial", 12), command=escanear_estandar).pack(pady=30)

# Etiqueta de estado
estado_label = tk.Label(ventana, text="", fg="blue", font=("Arial", 14))
estado_label.pack(pady=10)

# Título de la lista de estándares en uso
tk.Label(ventana, text="Estándares en uso:", font=("Arial", 16, "bold")).pack(pady=10)

# Crear Treeview
lista_estandares = ttk.Treeview(ventana, columns=('ID Producto', 'Código', 'Fecha Retiro', 'Empleado Retiro'), show='headings')

# Configurar encabezados
lista_estandares.heading('ID Producto', text='Estándar')
lista_estandares.heading('Código', text='Código')
lista_estandares.heading('Fecha Retiro', text='Fecha Retiro')
lista_estandares.heading('Empleado Retiro', text='Quimicos con Estandar en uso')

# Configurar ancho de columnas
lista_estandares.column('ID Producto', width=200, anchor='center')
lista_estandares.column('Código', width=200, anchor='center')
lista_estandares.column('Fecha Retiro', width=200, anchor='center')
lista_estandares.column('Empleado Retiro', width=200, anchor='center')

lista_estandares.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

# Inicializar la lista con los estándares actuales
actualizar_lista(lista_estandares)

# Ejecutar ventana
ventana.mainloop()
